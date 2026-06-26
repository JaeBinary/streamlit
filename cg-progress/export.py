"""Raw Data → 공식 양식(.xlsx) 채우기.

양식(form/*.xlsx)은 측정값 입력이 '한 열'만 있는 템플릿이다. Serial이 여러 개면 그 입력 열을
서식째(셀 스타일·열 너비·조건부 서식·데이터 검증) 우측으로 복제해 동일 포맷으로 채운다.
디스크 원본은 '읽기만' 하고(불변) openpyxl로 메모리 사본만 수정한다.

보드별 매핑(파일·시트·입력 열)은 constants.BOARD_CONFIG의 "form" 키에 있다.
행 좌표는 4종 양식 공통이라 여기 상수로 둔다.
"""

import io
from copy import copy
from datetime import datetime

import openpyxl
import streamlit as st
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.utils import column_index_from_string, get_column_letter

FORM_DIR = "form"

# 양식 공통 행 좌표 (1-base). 5행은 헤더, 6행부터 측정값(test_item 1 → 6행).
SERIAL_ROW, DATE_ROW, TESTER_ROW = 2, 3, 4
FIRST_DATA_ROW = 6


def _num(value):
    """측정값(DB TEXT)을 숫자면 float로, 아니면(예: 'Open') 원본 문자열로. 빈값은 None."""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def _ranges_in_column(sqref, col):
    """sqref(MultiCellRange) 중 지정 열에만 걸친 범위들의 (min_row, max_row) 목록."""
    out = []
    for rng in sqref.ranges:
        min_col, min_row, max_col, max_row = rng.bounds
        if min_col == col == max_col:
            out.append((min_row, max_row))
    return out


def _range_str(col, min_row, max_row):
    letter = get_column_letter(col)
    return f"{letter}{min_row}" if min_row == max_row else f"{letter}{min_row}:{letter}{max_row}"


def _clone_column(ws, src_col, dst_col, last_row):
    """src_col 열을 dst_col로 서식(+정적 값)·열 너비째 복제한다(2~last_row행)."""
    for row in range(SERIAL_ROW, last_row + 1):
        s = ws.cell(row, src_col)
        d = ws.cell(row, dst_col)
        d.value = s.value                 # 헤더 등 정적 값까지 동일하게(측정·헤더는 이후 덮어씀)
        if s.has_style:
            d._style = copy(s._style)     # font·fill·border·number_format·alignment 일괄 복제
    src_w = ws.column_dimensions[get_column_letter(src_col)].width
    if src_w is not None:
        ws.column_dimensions[get_column_letter(dst_col)].width = src_w


def _widen_conditional_formatting(ws, col0, last_col):
    """col0(입력 열)에만 걸린 조건부 서식의 '적용 대상'을 col0~last_col 전체로 넓힌다.

    열마다 규칙을 '복제'하면 우선순위 번호가 중복돼 Excel이 문서 순서로 재배치 →
    'Open'(True일 경우 중지)이 '범위 밖 빨강'보다 뒤로 밀려 빨강이 먼저 칠해진다.
    그래서 규칙 객체(고유 우선순위·중지 플래그·서식)를 그대로 둔 채 범위만 확장한다.
    수식 $D6,$E6는 열 고정이라 어느 열에서도 같은 행의 MIN/MAX와 비교돼 정확하다.
    """
    start, end = get_column_letter(col0), get_column_letter(last_col)
    entries = [(cf.sqref, list(cf.rules)) for cf in ws.conditional_formatting]
    ws.conditional_formatting = ConditionalFormattingList()
    for sqref, rules in entries:
        rows = _ranges_in_column(sqref, col0)
        # col0 범위는 col0~last_col로 확장, 그 외(있다면)는 원본 그대로 유지.
        target = " ".join(f"{start}{a}:{end}{b}" for a, b in rows) if rows else str(sqref)
        for rule in rules:
            ws.conditional_formatting.add(target, rule)


@st.cache_data(show_spinner=False)
def build_filled_form(form_file: str, form_sheet: str, serial_col: str,
                      prefix: str, n_steps: int, view) -> bytes:
    """검수완료·Serial 필터가 끝난 view를 공식 양식에 채워 xlsx bytes로 반환.

    view 컬럼: serial_number, test_item, measurements, test_datetime, test_By(이미 이름).
    Serial은 오름차순으로 입력 열부터 좌→우. 둘째 Serial부터는 입력 열을 서식·조건부 서식·
    데이터 검증까지 복제해 동일 포맷을 유지한다(설명/MIN/MAX·헤더·하단 결과 행은 그대로).
    @st.cache_data: st.tabs가 매 rerun에 모든 탭을 렌더해도 view가 그대로면 재생성하지 않는다.
    """
    wb = openpyxl.load_workbook(f"{FORM_DIR}/{form_file}")
    ws = wb[form_sheet]

    col0 = column_index_from_string(serial_col)
    last_row = ws.max_row
    serials = sorted(view["serial_number"].unique())

    # 입력 열(col0)에 걸린 데이터 검증(PASS/FAIL 드롭다운)을 미리 떠둔다(열 추가 전 기준).
    dv_src = [(dv, _ranges_in_column(dv.sqref, col0)) for dv in ws.data_validations.dataValidation]

    for offset, serial in enumerate(serials):
        col = col0 + offset
        if offset:
            # 둘째 Serial부터: 입력 열을 서식·열 너비째 복제하고 데이터 검증을 확장한다.
            _clone_column(ws, col0, col, last_row)
            for dv, ranges in dv_src:
                for min_row, max_row in ranges:
                    dv.add(_range_str(col, min_row, max_row))

        # Serial별 값 기록: 번호(접두사 뗀 숫자)·날짜·담당자·측정값.
        sub = view[view["serial_number"] == serial]
        first = sub.iloc[0]
        ws.cell(SERIAL_ROW, col).value = int(serial[len(prefix):])
        ws.cell(DATE_ROW, col).value = datetime.strptime(str(first["test_datetime"])[:10], "%Y-%m-%d")
        ws.cell(TESTER_ROW, col).value = first["test_By"]               # view에서 이미 oid→이름
        for _, rec in sub.iterrows():
            item = int(rec["test_item"])
            if 1 <= item <= n_steps:                                    # 양식 행 범위 안만 기록
                ws.cell(FIRST_DATA_ROW + item - 1, col).value = _num(rec["measurements"])

    # 조건부 서식은 복제 대신 '적용 대상' 범위를 추가 열까지 넓혀 우선순위 충돌을 피한다.
    if len(serials) > 1:
        _widen_conditional_formatting(ws, col0, col0 + len(serials) - 1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
