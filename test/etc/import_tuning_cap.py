"""Tuning Cap 테스트 시트(.xlsx) → cg_progress.db 1회성 이관 스크립트.

form/WG250-CG-Bypass and Tuning Capacitor Test sheet.xlsx 의 'Tuning Cap' 시트
(가로형: 열=Serial, 행=Test Item)를 읽어 기존 test_results 테이블에 적재한다.
기존 H-Bridge/Gate Drive/Bypass 데이터와의 키 충돌을 피하기 위해 serial은 'T'+4자리로
저장한다(시트 시리얼값 6~9 → T0006~T0009).

저장 포맷·SQL은 database.insert_records()와 동일하게 맞춘다(INSERT OR REPLACE).
T 접두어 + INSERT OR REPLACE라 재실행해도 중복 없이 갱신만 되어 멱등하다.
(import_gate_drive.py와 동일 로직. 차이점: 시트/경로/접두어, Test Item이 B열(2).)
"""
import sqlite3
from datetime import datetime
from pathlib import Path

import openpyxl

BASE = Path(__file__).parent
XLSX_PATH = BASE / "form" / "WG250-CG-Bypass and Tuning Capacitor Test sheet.xlsx"
DB_PATH = BASE / "data" / "cg_progress.db"
SHEET = "Tuning Cap"

SAVED_BY = "jaebin.kim@gwanakanalog.com"  # 현재 사용자
SERIAL_PREFIX = "T"                        # Tuning Capacitor 보드 구분 접두어

# 헤더 행 위치 (1-base)
ROW_SERIAL, ROW_DATE, ROW_TESTEDBY = 2, 3, 4
ROW_DATA_START = 6
COL_TEST_ITEM = 2                          # B열 (Gate Drive 시트는 C열이었음)
TEST_ITEM_FIX = {}                         # 보정할 입력 오타 없음


def _now() -> str:
    # database._now()와 동일 포맷
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _normalize_test_item(raw) -> str:
    """Test Item 셀값을 문자열로 정규화. 알려진 오타는 보정한다."""
    fixed = TEST_ITEM_FIX.get(raw, raw)
    if isinstance(fixed, float) and fixed.is_integer():
        fixed = int(fixed)
    return str(fixed)


def extract_rows():
    """엑셀에서 (serial, test_item, measurements, test_date, tested_by) 튜플 목록을 만든다."""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET]

    # 행 2에서 숫자 serial이 있는 컬럼만 측정값 컬럼으로 본다.
    serial_cols = {}  # col_idx -> serial 문자열(T0006)
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=ROW_SERIAL, column=c).value
        if isinstance(v, (int, float)):
            serial_cols[c] = f"{SERIAL_PREFIX}{int(v):04d}"

    # 저장 순서를 serial 우선으로 한다: T0006의 test_item 1~ → T0007의 1~ → …
    # (바깥 루프=serial, 안쪽 루프=데이터 행)
    rows = []
    for c, serial in serial_cols.items():
        date_cell = ws.cell(row=ROW_DATE, column=c).value
        test_date = date_cell.date().isoformat() if isinstance(date_cell, datetime) else str(date_cell)
        tested_by = ws.cell(row=ROW_TESTEDBY, column=c).value

        for r in range(ROW_DATA_START, ws.max_row + 1):
            item_raw = ws.cell(row=r, column=COL_TEST_ITEM).value
            if item_raw is None:
                continue  # 빈 행 / Overall Result 행 등은 건너뜀
            test_item = _normalize_test_item(item_raw)

            meas = ws.cell(row=r, column=c).value
            if meas is None or (isinstance(meas, str) and meas.strip() == ""):
                continue  # 측정값 없는 칸은 행 생성 안 함

            rows.append((serial, test_item, str(meas), test_date, str(tested_by)))
    return rows


def main():
    rows = extract_rows()
    now = _now()
    payload = [
        (serial, test_item, meas, test_date, tested_by, now, SAVED_BY)
        for serial, test_item, meas, test_date, tested_by in rows
    ]

    conn = sqlite3.connect(DB_PATH)
    try:
        with conn:
            # serial 우선 순서로 다시 쌓기 위해 기존 T 행을 먼저 비운다(저장 순서 보장).
            conn.execute(f"DELETE FROM test_results WHERE serial LIKE '{SERIAL_PREFIX}%'")
            conn.executemany(
                "INSERT OR REPLACE INTO test_results"
                " (serial, test_item, measurements, test_date, tested_by, save_datetime, saved_by)"
                " VALUES (?,?,?,?,?,?,?)",
                payload,
            )
    finally:
        conn.close()

    serials = sorted({s for s, *_ in rows})
    items = sorted({int(ti) for _, ti, *_ in rows})
    print(f"이관 완료: {len(payload)}건")
    print(f"  Serial({len(serials)}개): {', '.join(serials)}")
    print(f"  Test Item({len(items)}개): {items[0]}~{items[-1]}")


if __name__ == "__main__":
    main()
